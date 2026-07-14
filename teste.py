import pandas as pd
from openpyxl import load_workbook, xl

wb = load_workbook("your_file.xlsx")
ws = wb.active

data = ws["A1:C4"]
df = pd.DataFrame(data, columns=["Column1", "Column2", "Column3"])